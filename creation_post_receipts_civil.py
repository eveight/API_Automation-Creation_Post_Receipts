import logging
import time

from database import Session
from exceptions.pages_exceptions import CreateReceiptErrorException
from exceptions.pages_exceptions import EmptyDBDataException
from models import Receipt, PaidStatusEnum
from datetime import datetime
import requests
import os

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from config.modal_log_handler import ModalLogHandler
from .base import GetDataDB

from .base import Mixin

from .base import read_files_in_folder
from models import StatusSudEnum

logger = logging.getLogger(__name__)


class CreateReceiptException(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


class UpdateReceiptsStatus:
    def __init__(self, user_id, type_of_receipt):
        self.data_block = None
        self.user_id = user_id
        self.type_of_receipt = type_of_receipt

    def update_status_in_db(self, data_block):
        self.data_block = data_block
        for v in data_block:
            session = Session()
            try:
                pinfl_of_debtor = v.get('pinfl_of_debtor')
                receipt = session.query(Receipt).filter_by(pinfl_of_debtor=pinfl_of_debtor,
                                                           status_of_created=False,
                                                           type_of_sud=self.type_of_receipt,
                                                           user_id=self.user_id).first()
                if receipt:
                    receipt.status_of_created = True
                    receipt.number_of_receipt = v.get('invoice')
                    receipt.sum_of_receipt = v.get('sum_of_receipt')
                    receipt.link_of_receipt = v.get('link_of_receipt')
                    receipt.paid = PaidStatusEnum.CREATED
                    receipt.receipt_created_at = datetime.now()
                    session.commit()
                else:
                    logger.error(f"Квитанция с ПИНФЛ {pinfl_of_debtor} не найдена или уже была создана")
                    raise Exception(f"Квитанция с ПИНФЛ {pinfl_of_debtor} не найдена или уже была создана.")
            except Exception as e:
                logger.error(f"Ошибка обновления строки в БД: {e}",
                             extra={"user_message": "Ошибка обновления строки в БД."})
                session.rollback()
            finally:
                session.close()


class CreateReceiptAPI(Mixin):

    def __init__(self, data, user_id, type_of_receipt):
        self.data = data
        self.user_id = user_id
        self.type_of_receipt = type_of_receipt

        self.type_name_of_output()

    def type_name_of_output(self):
        if self.type_of_receipt == StatusSudEnum.CIVIL:
            self.name_of_created_excel = 'created_receipts_info.xlsx'
            self.name_of_not_created_excel = 'not_created_receipts_info.xlsx'
        else:
            self.name_of_created_excel = 'created_receipts_info_post.xlsx'
            self.name_of_not_created_excel = 'not_created_receipts_info_post.xlsx'

    def request_create_receipt(self, court_id, pinfl_of_debtor, name_of_client, stir_of_client, address_of_client, sum_of_debt):
        time.sleep(2)
        headers = self.HEADERS.copy()
        headers['Referer'] = 'https://***/create-receipt'
        headers['Origin'] = 'https://***'
        headers['Content-Type'] = 'application/json'
        headers['Content-Length'] = '371'

        body = {
            "entityType": "JURIDICAL",
            "juridicalEntity": {
                "name": name_of_client,
                "tin": stir_of_client,
                "address": address_of_client,
            },
            "amount": sum_of_debt * 100,
            "overdue": 0,
            "courtType": "CITIZEN",
            "courtId": court_id,
            "description": "",
            "isInFavor": True,
            "purposeId": 1,
        }

        if self.type_of_receipt == StatusSudEnum.CIVIL:
            body["payCategoryId"] = 1
        else:
            body["payCategoryId"] = 3

        response = requests.post('https://***/api/invoice/create', json=body, headers=headers)
        response.raise_for_status()
        if response.status_code == 201:
            logger.info(f"Квитанция для ПИНФЛ {pinfl_of_debtor} успешно создана")
            return response.json()
        else:
            logger.error(f"Статус код не равен 201 при создании квитанции,\nОтвет от сервера: {response.json()}",
                         extra={"user_message": "Ошибка отправки запроса на создание квитанции"})
            raise CreateReceiptException(f"Статус код не равен 201 при создании квитанции: {response.json()}")

    def _write_to_excel(self, data, filename):
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            headers = ['ФИО', 'Пинфл должника', 'Номер квитанции', 'Название региона', 'Сумма квитанции', 'Ссылка на квитанцию']
            ws.append(headers)

        ws.append(data)
        wb.save(filename)

    def _create_receipts(self, data_block):
        for k, v in data_block.items():
            try:
                fio = v.get('fio')
                pinfl_of_debtor = v.get('pinfl_of_debtor')
                name_of_client = v.get('name_of_client')
                stir_of_client = v.get('stir_of_client')
                address_of_client = v.get('address_of_client')
                sum_of_debt = v.get('sum_of_debt')
                cort_id = v.get('id_sud')
                if cort_id:
                    response = self.request_create_receipt(cort_id, pinfl_of_debtor, name_of_client,
                                                           stir_of_client, address_of_client, sum_of_debt)
                    invoice = response.get('invoice')
                    v['invoice'] = invoice
                    v['sum_of_receipt'] = sum_of_debt
                    v['link_of_receipt'] = f'https://***/invoice/{invoice}'
                    UpdateReceiptsStatus(self.user_id, self.type_of_receipt).update_status_in_db([v])
                else:
                    logger.error(f'Не удалось получить регион для строки {k} - {v}')
                    raise CreateReceiptException(f'Не удалось получить регион для строки {k} - {v}')
            except CreateReceiptException as e:
                logger.error(f"Ошибка при создании квитанции,\nОшибка: {e}",
                             extra={"user_message": f'Квитанция для {k} - {v} не создана: {e}'})
            except Exception as e:
                logger.error(f"Ошибка при создании квитанции для {k} - {v},\nОшибка: {e}",
                             extra={"user_message": f'Что-то пошло не так с {k} - {v}.'})

    def process_create_receipts(self):
        data_for_blocking = self.data

        block_size = 10
        data_blocks = [dict(list(data_for_blocking.items())[i:i + block_size]) for i in
                       range(0, len(data_for_blocking), block_size)]

        for block in data_blocks:
            self._create_receipts(block)


def run(user_id, type_of_receipt, modal_window):
    modal_handler = ModalLogHandler()
    modal_handler.set_modal_window(modal_window)
    logger.addHandler(modal_handler)
    try:
        data_for_automation = GetDataDB(user_id, type_of_receipt).get_data_from_db_to_create_receipts()
        if not data_for_automation:
            logger.info('Нет доступных данных для создание квитанций')
            raise EmptyDBDataException()
        obj_create_receipt_api = CreateReceiptAPI(data=data_for_automation, user_id=user_id, type_of_receipt=type_of_receipt)
        obj_create_receipt_api.process_create_receipts()

        logger.removeHandler(modal_handler)
    except Exception as e:
        if isinstance(e, EmptyDBDataException):
            logger.info(f'Обнаружены проблемы с созданием квитанций: {e}')
        else:
            logger.error(f"Ошибка при создании квитанций: {e}",
                         extra={'user_message': f'Обнаружены проблемы с созданием квитанций'})
        raise CreateReceiptErrorException()
    finally:
        logger.removeHandler(modal_handler)
