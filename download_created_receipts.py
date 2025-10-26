import asyncio
import logging

import aiofiles
import aiohttp
from aiofiles import ospath
from aiofiles import os as os_aio

import os
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from exceptions.pages_exceptions import DownloadReceiptErrorException
from exceptions.pages_exceptions import EmptyDBDataException
from config.modal_log_handler import ModalLogHandler
from models import PaidStatusEnum
from .base import GetDataDB, Mixin
from .base import run_check_paid_receipt
from models import StatusSudEnum

logger = logging.getLogger(__name__)


class DownloadReceiptException(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


class DownloadCreatedReceipt(Mixin):
    def __init__(self, user_id, type_of_receipt, path: str):
        self.full_path = path
        self.user_id = user_id
        self.type_of_receipt = type_of_receipt

        self.type_name_of_output()

    def type_name_of_output(self):
        if self.type_of_receipt == StatusSudEnum.CIVIL:
            self.name_of_not_created_excel = 'not_created_receipts_info.xlsx'
            self.prefix_for_file = 'DB '
        else:
            self.name_of_not_created_excel = 'not_created_receipts_info_post.xlsx'
            self.prefix_for_file = 'PR '

    def _write_to_excel(self, data, filename):
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            headers = ['Пинфл должника', 'ФИО', 'Номер квитанции', 'Название региона', 'Сумма квитанции', 'Ссылка на квитанцию']
            ws.append(headers)

        ws.append(data)
        wb.save(filename)

    async def async_request_download_pdf(self, session, semaphore, invoice_number, filename, pinfl):
        _dir_name = filename
        if not await ospath.isdir(os.path.join(self.full_path, _dir_name)):
            _dir_name = f"auto_create_{filename}"
            try:
                await os_aio.mkdir(os.path.join(self.full_path, _dir_name))
            except FileExistsError as e:
                print(f"FILE EXIST == {e}")

        full_path = os.path.join(
            self.full_path,
            _dir_name,
            f'{self.prefix_for_file} {filename}_{pinfl}_{invoice_number}.pdf'
        )

        headers = self.HEADERS.copy()
        headers['Referer'] = f'https://***/invoice/{invoice_number}'

        params = {"invoice": invoice_number}
        try:
            async with semaphore:
                logger.info(f'Выполняется запрос для загрузки файла {filename} {invoice_number}')
                async with session.get('https://***/api/invoice/asDocument', params=params,
                                       headers=headers, timeout=120) as response:
                    response.raise_for_status()
                    if response.status == 200:
                        async with aiofiles.open(full_path, 'wb') as f:
                            await f.write(await response.read())
                        logger.info(f"Квитанция сохранена по пути: {full_path}")
                    else:
                        logger.error(f"Статус ответа от сервера не равен 200 при загрузке документа {invoice_number},\nСтатус: {response.status}",
                                     extra={"user_message": f"Ошибка при загрузке файла: {invoice_number}"})
        except asyncio.TimeoutError as e:
            logger.error(f'Превышен таймаут при загрузке файла')
        except Exception as e:
            logger.error(f"Ошибка при попытке скачать файл {invoice_number},\nОшибка: {e}",
                         extra={'user_message': f"Ошибка при загрузке файла {invoice_number}"})

        logger.info(f'Скачивание pdf файла для {filename} {pinfl} {invoice_number} завершено')

    async def _download_pdf(self, session, data_block):
        semaphore = asyncio.Semaphore(10)
        tasks = []
        for fio, pinfl, invoice in data_block:
            try:
                task = self.async_request_download_pdf(session, semaphore, invoice, fio, pinfl)
                tasks.append(task)
            except Exception as e:
                logger.error(f"Ошибка при загрузке квитанции {invoice},\nОшибка: {e}",
                             extra={"user_message": f'Что-то пошло не так с {invoice} - {fio} - {pinfl}.'})
        await asyncio.gather(*tasks)

    async def async_process_for_download_pdf(self, ids):
        data_for_blocking = GetDataDB(self.user_id, self.type_of_receipt).get_data_from_db_for_download_pdf_today(ids)

        block_size = 10
        data_blocks = [data_for_blocking[i:i + block_size] for i in range(0, len(data_for_blocking), block_size)]
        async with aiohttp.ClientSession() as session:
            for block in data_blocks:
                await self._download_pdf(session, block)


def run(user_id, type_of_receipt, receipt_path, modal_window, ids):
    modal_handler = ModalLogHandler()
    modal_handler.set_modal_window(modal_window)
    logger.addHandler(modal_handler)

    try:
        run_check_paid_receipt(user_id, type_of_receipt, ids)
        data_for_automation = GetDataDB(user_id, type_of_receipt).get_data_from_db_for_download_pdf_today(ids, check=True)
        if not data_for_automation:
            logger.info('Нет доступных квитанций для скачивания.')
            raise EmptyDBDataException('Нет созданных квитанций для скачивания')

        for index, receipt in enumerate(data_for_automation):
            if receipt[1] != PaidStatusEnum.PAID:
                params = (
                    ("\"\"", "не создана")
                    if receipt[-1] is None else
                    (receipt[-1], "не оплачена")
                )
                logger.info(
                    f"Квитанция {params[0]}, ФИО - {receipt[0]} не скачана так как {params[1]}."
                )

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        obj_create_receipt_api = DownloadCreatedReceipt(user_id, type_of_receipt, receipt_path)
        loop.run_until_complete(obj_create_receipt_api.async_process_for_download_pdf(ids))
    except Exception as e:
        if isinstance(e, EmptyDBDataException):
            logger.info(f'Обнаружены проблемы со скачиванием квитанций: {e}')
        else:
            logger.error(f"Ошибка при загрузке квитанций: {e}",
                         extra={'user_message': f'Обнаружены проблемы со скачиванием квитанций'})
        raise e
    finally:
        logger.removeHandler(modal_handler)

